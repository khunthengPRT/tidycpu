#!/usr/bin/env python3
"""
TidyCPU — CPU Affinity Optimization Utility
============================================
Balances CPU load by reassigning process affinity across cores.
Target: Linux x64 (Debian/Ubuntu)
Requires: root privileges
"""

import os
import sys
import re
import subprocess
import time
import argparse
from dataclasses import dataclass, field
from typing import Optional
from collections import defaultdict
from datetime import datetime

# ─────────────────────────────────────────────
# ANSI Color Palette
# ─────────────────────────────────────────────
class C:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    RED     = "\033[91m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    BLUE    = "\033[94m"
    MAGENTA = "\033[95m"
    CYAN    = "\033[96m"
    WHITE   = "\033[97m"
    BG_RED  = "\033[41m"
    BG_DARK = "\033[100m"

# ─────────────────────────────────────────────
# Data Structures
# ─────────────────────────────────────────────
@dataclass
class CPUTopology:
    """Physical CPU topology info."""
    physical_id:  int
    core_id:      int
    logical_id:   int
    is_hyperthread: bool = False

@dataclass
class SystemInfo:
    """System-wide hardware information."""
    cpu_model:     str
    total_memory:  str  # e.g., "15.6 GiB"
    available_memory: str
    kernel_cmdline: str
    cpu_freq_min:  Optional[float] = None  # MHz
    cpu_freq_max:  Optional[float] = None  # MHz
    cpu_freq_cur:  Optional[float] = None  # MHz

@dataclass
class CoreStat:
    core_id:  int
    usage:    float   # 0.0 – 100.0 percent
    label:    str     # HOT / WARM / COLD
    pids:     list    = field(default_factory=list)
    physical_id: Optional[int] = None
    core_within_physical: Optional[int] = None
    top_proc: str     = ""  # name of the busiest process/thread on this core
    top_parent: str   = ""  # name of the parent process owning that thread
    all_procs: list   = field(default_factory=list)  # [(thread_name, parent_name, cpu_pct), ...]

@dataclass
class ThreadInfo:
    """Individual thread details."""
    tid:          int
    tgid:         int  # parent process group id
    name:         str
    cpu_percent:  float
    current_cores: list[int]

@dataclass
class ProcessInfo:
    pid:          int
    name:         str
    cpu_percent:  float
    current_cores: list[int]
    affinity_mask: str
    threads:      list[ThreadInfo] = field(default_factory=list)

@dataclass
class RebalanceAction:
    pid:           int
    name:          str
    cpu_percent:   float
    from_cores:    list[int]
    to_cores:      list[int]
    manual_only:   bool = False
    error_msg:     str  = ""

@dataclass
class Snapshot:
    """Single snapshot of system state during live monitoring."""
    timestamp:    str
    iteration:    int
    core_stats:   list[CoreStat]
    processes:    list[ProcessInfo]

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def die(msg: str):
    print(f"\n{C.RED}{C.BOLD}[FATAL]{C.RESET} {msg}\n")
    sys.exit(1)

def run(cmd: list[str], timeout: int = 5) -> tuple[int, str, str]:
    """Run a subprocess, return (returncode, stdout, stderr)."""
    try:
        r = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout
        )
        return r.returncode, r.stdout.strip(), r.stderr.strip()
    except subprocess.TimeoutExpired:
        return -1, "", "Command timed out"
    except FileNotFoundError:
        return -1, "", f"Command not found: {cmd[0]}"

def parse_cpulist(cpulist_str: str) -> list[int]:
    """Parse kernel cpulist format '0,2-5,7' → [0, 2, 3, 4, 5, 7]."""
    cores = []
    for part in cpulist_str.split(","):
        part = part.strip()
        if "-" in part:
            lo, hi = part.split("-")
            cores.extend(range(int(lo), int(hi) + 1))
        elif part.isdigit():
            cores.append(int(part))
    return sorted(set(cores))

def cores_to_cpulist(cores: list[int]) -> str:
    """[0, 2, 3, 4] → '0,2-4'."""
    if not cores:
        return ""
    cores = sorted(set(cores))
    parts, start, end = [], cores[0], cores[0]
    for c in cores[1:]:
        if c == end + 1:
            end = c
        else:
            parts.append(str(start) if start == end else f"{start}-{end}")
            start = end = c
    parts.append(str(start) if start == end else f"{start}-{end}")
    return ",".join(parts)

def mask_to_cores(hex_mask: str, num_cores: int) -> list[int]:
    """Convert hex affinity mask to list of core IDs."""
    try:
        mask = int(hex_mask.replace(",", ""), 16)
        return [i for i in range(num_cores) if mask & (1 << i)]
    except ValueError:
        return []

def resolve_pid(pid_or_name: str) -> int:
    """
    Resolve a single process name or PID string to an integer PID.
    If the value is numeric, return it directly.
    Otherwise, use pgrep to find the PID by process name.
    Returns the first matched PID, or exits with an error.
    """
    if pid_or_name.isdigit():
        return int(pid_or_name)

    # Name-based lookup via pgrep (exact match first)
    rc, out, _ = run(["pgrep", "-x", pid_or_name])
    if rc == 0 and out:
        pids = out.splitlines()
        if len(pids) > 1:
            print(f"  {C.YELLOW}⚠  Multiple processes match '{pid_or_name}': {', '.join(pids)}{C.RESET}")
            print(f"  {C.YELLOW}   Using PID {pids[0]}{C.RESET}")
        return int(pids[0])

    # Fallback: partial match
    rc2, out2, _ = run(["pgrep", pid_or_name])
    if rc2 == 0 and out2:
        pids = out2.splitlines()
        if len(pids) > 1:
            print(f"  {C.YELLOW}⚠  Multiple processes match '{pid_or_name}': {', '.join(pids)}{C.RESET}")
            print(f"  {C.YELLOW}   Using PID {pids[0]}{C.RESET}")
        return int(pids[0])

    die(f"No process found matching name '{pid_or_name}'.\n"
        f"  Try: pgrep {pid_or_name}  to verify the process exists.")

def resolve_pids(pid_spec: str) -> list[int]:
    """
    Resolve a pipe-separated list of PIDs/names to a deduplicated list of integer PIDs.
    Examples:
      "1234"           → [1234]
      "nginx"          → [<pid of nginx>]
      "nginx|php-fpm|1234" → [<nginx pid>, <php-fpm pid>, 1234]
    """
    tokens = [t.strip() for t in pid_spec.split("|") if t.strip()]
    seen: set[int] = set()
    result: list[int] = []
    for token in tokens:
        pid = resolve_pid(token)
        if pid not in seen:
            seen.add(pid)
            result.append(pid)
    return result

# ─────────────────────────────────────────────
# Step 0 – Privilege Check
# ─────────────────────────────────────────────
def check_root():
    if os.geteuid() != 0:
        die(
            "TidyCPU requires root privileges.\n"
            f"  {C.YELLOW}Re-run with: sudo python3 tidycpu.py{C.RESET}"
        )

# ─────────────────────────────────────────────
# System Information Collection
# ─────────────────────────────────────────────
def get_system_info(show_cpu_freq: bool = False) -> SystemInfo:
    """
    Collect system-wide hardware information.
    """
    # CPU Model from /proc/cpuinfo
    cpu_model = "Unknown"
    try:
        with open("/proc/cpuinfo") as f:
            for line in f:
                if line.startswith("model name"):
                    cpu_model = line.split(":", 1)[1].strip()
                    break
    except FileNotFoundError:
        pass
    
    # Memory info from /proc/meminfo
    total_mem = "Unknown"
    avail_mem = "Unknown"
    try:
        with open("/proc/meminfo") as f:
            meminfo = {}
            for line in f:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    key = parts[0].strip()
                    value = parts[1].strip().split()[0]  # Get numeric part
                    meminfo[key] = int(value)
            
            if "MemTotal" in meminfo:
                total_kb = meminfo["MemTotal"]
                total_mem = f"{total_kb / 1024 / 1024:.1f} GiB"
            
            if "MemAvailable" in meminfo:
                avail_kb = meminfo["MemAvailable"]
                avail_mem = f"{avail_kb / 1024 / 1024:.1f} GiB"
    except (FileNotFoundError, ValueError):
        pass
    
    # Kernel command line
    cmdline = "Unknown"
    try:
        with open("/proc/cmdline") as f:
            cmdline = f.read().strip()
    except FileNotFoundError:
        pass
    
    # CPU frequency info (optional)
    freq_min, freq_max, freq_cur = None, None, None
    if show_cpu_freq:
        try:
            # Try cpufreq interface for CPU0
            with open("/sys/devices/system/cpu/cpu0/cpufreq/cpuinfo_min_freq") as f:
                freq_min = int(f.read().strip()) / 1000  # Convert kHz to MHz
            with open("/sys/devices/system/cpu/cpu0/cpufreq/cpuinfo_max_freq") as f:
                freq_max = int(f.read().strip()) / 1000
            with open("/sys/devices/system/cpu/cpu0/cpufreq/scaling_cur_freq") as f:
                freq_cur = int(f.read().strip()) / 1000
        except (FileNotFoundError, ValueError, PermissionError):
            # Fallback: try to get from cpuinfo
            try:
                with open("/proc/cpuinfo") as f:
                    for line in f:
                        if line.startswith("cpu MHz"):
                            freq_cur = float(line.split(":", 1)[1].strip())
                            break
            except (FileNotFoundError, ValueError):
                pass
    
    return SystemInfo(
        cpu_model=cpu_model,
        total_memory=total_mem,
        available_memory=avail_mem,
        kernel_cmdline=cmdline,
        cpu_freq_min=freq_min,
        cpu_freq_max=freq_max,
        cpu_freq_cur=freq_cur
    )

# ─────────────────────────────────────────────
# CPU Topology Detection
# ─────────────────────────────────────────────
def get_cpu_topology() -> dict[int, CPUTopology]:
    """
    Parse /sys/devices/system/cpu/cpu*/topology to map logical cores
    to physical CPUs and physical cores.
    """
    topology = {}
    cpu_dirs = sorted(
        [d for d in os.listdir("/sys/devices/system/cpu") if d.startswith("cpu") and d[3:].isdigit()],
        key=lambda x: int(x[3:])
    )
    
    for cpu_dir in cpu_dirs:
        logical_id = int(cpu_dir[3:])
        base = f"/sys/devices/system/cpu/{cpu_dir}/topology"
        
        try:
            with open(f"{base}/physical_package_id") as f:
                physical_id = int(f.read().strip())
            with open(f"{base}/core_id") as f:
                core_id = int(f.read().strip())
            
            # Check if this is a hyperthread sibling
            with open(f"{base}/thread_siblings_list") as f:
                siblings = parse_cpulist(f.read().strip())
                is_hyperthread = len(siblings) > 1 and logical_id != min(siblings)
            
            topology[logical_id] = CPUTopology(
                physical_id=physical_id,
                core_id=core_id,
                logical_id=logical_id,
                is_hyperthread=is_hyperthread
            )
        except (FileNotFoundError, ValueError):
            # Fallback for systems without topology info
            topology[logical_id] = CPUTopology(
                physical_id=0,
                core_id=logical_id,
                logical_id=logical_id,
                is_hyperthread=False
            )
    
    return topology

# ─────────────────────────────────────────────
# Step 1 – Telemetry: Read /proc/stat (two snapshots)
# ─────────────────────────────────────────────
def read_proc_stat() -> dict[int, dict]:
    """Parse /proc/stat and return per-core raw counters."""
    cores = {}
    try:
        with open("/proc/stat") as f:
            for line in f:
                m = re.match(r"^cpu(\d+)\s+(.+)$", line)
                if m:
                    cid = int(m.group(1))
                    vals = list(map(int, m.group(2).split()))
                    # user nice system idle iowait irq softirq steal guest guest_nice
                    cores[cid] = {
                        "user":    vals[0], "nice":    vals[1],
                        "system":  vals[2], "idle":    vals[3],
                        "iowait":  vals[4] if len(vals) > 4 else 0,
                        "irq":     vals[5] if len(vals) > 5 else 0,
                        "softirq": vals[6] if len(vals) > 6 else 0,
                        "steal":   vals[7] if len(vals) > 7 else 0,
                    }
    except FileNotFoundError:
        die("/proc/stat not found. Are you on Linux?")
    return cores

def get_top_proc_per_core() -> dict[int, tuple[str, str]]:
    """
    Read /proc/*/stat and /proc/*/task/*/stat to find the process or thread
    with the highest recent CPU time on each logical core.
    Returns a dict mapping core_id -> (thread_name, parent_process_name).
    """
    # Maps core_id -> (max_cputime, thread_name, parent_name)
    best: dict[int, tuple[int, str, str]] = {}

    try:
        pids = [p for p in os.listdir("/proc") if p.isdigit()]
    except PermissionError:
        return {}

    for pid_str in pids:
        pid = int(pid_str)
        task_dir = f"/proc/{pid}/task"

        # Read parent process name once per PID
        try:
            with open(f"/proc/{pid}/comm") as f:
                parent_name = f.read().strip()[:15]
        except (FileNotFoundError, PermissionError):
            parent_name = ""

        try:
            tids = [t for t in os.listdir(task_dir) if t.isdigit()]
        except (FileNotFoundError, PermissionError):
            continue

        for tid_str in tids:
            try:
                with open(f"{task_dir}/{tid_str}/stat") as f:
                    raw = f.read()
                # stat format: pid (comm) state ppid pgroup session tty_nr
                #   tpgid flags minflt cminflt majflt cmajflt
                #   utime stime cutime cstime priority nice ...
                #   (38) last_cpu
                # comm may contain spaces so we parse past the closing ')'
                rp = raw.rfind(")")
                if rp == -1:
                    continue
                fields = raw[rp + 2:].split()
                # fields[0] = state, fields[11]=utime, fields[12]=stime, fields[36]=processor
                if len(fields) < 37:
                    continue
                utime    = int(fields[11])
                stime    = int(fields[12])
                cpu_time = utime + stime
                last_cpu = int(fields[36])

                # Read thread comm for the name
                comm_path = f"{task_dir}/{tid_str}/comm"
                with open(comm_path) as f:
                    name = f.read().strip()[:15]

                cur_best = best.get(last_cpu)
                if cur_best is None or cpu_time > cur_best[0]:
                    best[last_cpu] = (cpu_time, name, parent_name)
            except (FileNotFoundError, PermissionError, ValueError, IndexError):
                continue

    return {core: (name, parent) for core, (_, name, parent) in best.items()}


def get_all_procs_per_core(
    min_usage: float = 0.0,
    ignore_procs: Optional[list[str]] = None,
) -> dict[int, list[tuple]]:
    """
    Scan all threads and group by the last CPU core they ran on.
    Returns dict: core_id -> [(thread_name, parent_name, cpu_pct), ...] sorted desc by cpu_pct.
    min_usage: skip threads with cpu_pct below this threshold.
    ignore_procs: list of name prefixes to exclude (case-insensitive prefix match).
    """
    ignore_prefixes = tuple(p.lower() for p in (ignore_procs or []))

    # Collect per-thread CPU% from ps (thread-level via -L)
    rc, out, _ = run(["ps", "-eLo", "tid,pcpu,comm", "--no-headers"])
    ps_map: dict[int, float] = {}  # tid -> cpu_pct
    if rc == 0:
        for line in out.splitlines():
            parts = line.split(None, 2)
            if len(parts) >= 2:
                try:
                    ps_map[int(parts[0])] = float(parts[1])
                except ValueError:
                    pass

    result: dict[int, list] = {}

    try:
        pids = [p for p in os.listdir("/proc") if p.isdigit()]
    except PermissionError:
        return {}

    for pid_str in pids:
        pid = int(pid_str)
        task_dir = f"/proc/{pid}/task"

        try:
            with open(f"/proc/{pid}/comm") as f:
                parent_name = f.read().strip()[:15]
        except (FileNotFoundError, PermissionError):
            parent_name = ""

        if ignore_prefixes and parent_name.lower().startswith(ignore_prefixes):
            continue

        try:
            tids = [t for t in os.listdir(task_dir) if t.isdigit()]
        except (FileNotFoundError, PermissionError):
            continue

        for tid_str in tids:
            tid = int(tid_str)
            try:
                with open(f"{task_dir}/{tid_str}/stat") as f:
                    raw = f.read()
                rp = raw.rfind(")")
                if rp == -1:
                    continue
                fields = raw[rp + 2:].split()
                if len(fields) < 37:
                    continue
                last_cpu = int(fields[36])

                with open(f"{task_dir}/{tid_str}/comm") as f:
                    thread_name = f.read().strip()[:15]

                if ignore_prefixes and thread_name.lower().startswith(ignore_prefixes):
                    continue

                cpu_pct = ps_map.get(tid, 0.0)
                if cpu_pct < min_usage:
                    continue

                result.setdefault(last_cpu, []).append((thread_name, parent_name, cpu_pct))
            except (FileNotFoundError, PermissionError, ValueError, IndexError):
                continue

    for core_id in result:
        result[core_id].sort(key=lambda x: x[2], reverse=True)

    return result


def get_core_usage(sample_ms: int = 500, topology: Optional[dict] = None,
                   stack_procs: bool = False, min_usage: float = 0.0,
                   ignore_procs: Optional[list[str]] = None) -> list[CoreStat]:
    """Two-snapshot delta to calculate real per-core CPU %."""
    snap1 = read_proc_stat()
    time.sleep(sample_ms / 1000)
    snap2 = read_proc_stat()

    stats = []
    for cid in sorted(snap1.keys()):
        s1, s2 = snap1[cid], snap2[cid]

        idle1  = s1["idle"] + s1["iowait"]
        idle2  = s2["idle"] + s2["iowait"]
        total1 = sum(s1.values())
        total2 = sum(s2.values())

        d_idle  = idle2  - idle1
        d_total = total2 - total1

        usage = 0.0 if d_total == 0 else (1.0 - d_idle / d_total) * 100.0

        if   usage >= 80: label = "HOT"
        elif usage >= 40: label = "WARM"
        else:             label = "COLD"

        phys_id = None
        core_in_phys = None
        if topology and cid in topology:
            phys_id = topology[cid].physical_id
            core_in_phys = topology[cid].core_id

        stats.append(CoreStat(
            core_id=cid,
            usage=round(usage, 1),
            label=label,
            physical_id=phys_id,
            core_within_physical=core_in_phys
        ))

    # Populate top_proc and top_parent from /proc after both snapshots are done
    top_procs = get_top_proc_per_core()
    for cs in stats:
        cs.top_proc, cs.top_parent = top_procs.get(cs.core_id, ("", ""))

    # Optionally populate all_procs for stacked display
    if stack_procs:
        all_procs_map = get_all_procs_per_core(min_usage=min_usage, ignore_procs=ignore_procs)
        for cs in stats:
            cs.all_procs = all_procs_map.get(cs.core_id, [])

    return stats

# ─────────────────────────────────────────────
# Step 2 – Analysis: Top CPU Processes + Affinity
# ─────────────────────────────────────────────
def get_threads_for_pid(pid: int, num_cores: int = 1) -> list[ThreadInfo]:
    """Get all threads for a specific PID with their CPU usage."""
    threads = []
    try:
        # Read thread list from /proc/<pid>/task/
        task_dir = f"/proc/{pid}/task"
        if not os.path.exists(task_dir):
            return threads
        
        for tid_str in os.listdir(task_dir):
            if not tid_str.isdigit():
                continue
            tid = int(tid_str)
            
            try:
                # Get thread name
                with open(f"{task_dir}/{tid}/comm") as f:
                    name = f.read().strip()[:24]
                
                # Get thread CPU usage from /proc/<pid>/task/<tid>/stat
                with open(f"{task_dir}/{tid}/stat") as f:
                    stat = f.read().strip()
                    pass
                
                # Use ps to get thread-level CPU
                rc, out, _ = run(["ps", "-p", str(tid), "-o", "pcpu", "--no-headers"])
                cpu = 0.0
                if rc == 0 and out:
                    try:
                        cpu = float(out.strip())
                    except ValueError:
                        pass
                
                # Get thread affinity
                rc2, out2, _ = run(["taskset", "-p", str(tid)])
                cur_cores = []
                if rc2 == 0:
                    m = re.search(r"current affinity mask:\s*([0-9a-fA-F,]+)", out2)
                    if m:
                        cur_cores = mask_to_cores(m.group(1), num_cores)
                
                threads.append(ThreadInfo(
                    tid=tid,
                    tgid=pid,
                    name=name,
                    cpu_percent=cpu,
                    current_cores=cur_cores
                ))
            except (FileNotFoundError, PermissionError):
                continue
    except (FileNotFoundError, PermissionError):
        pass
    
    return threads

def get_top_processes(n: int = 5, num_cores: int = 1, with_threads: bool = False) -> list[ProcessInfo]:
    """Get top-N CPU-consuming processes with their core affinity."""
    rc, out, err = run([
        "ps", "ax", "-o", "pid,pcpu,comm", "--sort=-pcpu", "--no-headers"
    ])
    if rc != 0:
        die(f"ps failed: {err}")

    procs = []
    for line in out.splitlines()[:20]:
        parts = line.split(None, 2)
        if len(parts) < 3:
            continue
        try:
            pid  = int(parts[0])
            cpu  = float(parts[1])
            name = parts[2].strip()[:24]
        except ValueError:
            continue

        # Get affinity mask via taskset
        rc2, out2, err2 = run(["taskset", "-p", str(pid)])
        if rc2 != 0:
            affinity_mask = "N/A"
            cur_cores     = []
        else:
            m = re.search(r"current affinity mask:\s*([0-9a-fA-F,]+)", out2)
            affinity_mask = m.group(1) if m else "N/A"
            cur_cores     = mask_to_cores(affinity_mask, num_cores) if m else []

        threads = []
        if with_threads:
            threads = get_threads_for_pid(pid, num_cores)

        procs.append(ProcessInfo(
            pid=pid, name=name, cpu_percent=cpu,
            current_cores=cur_cores, affinity_mask=affinity_mask,
            threads=threads
        ))
        if len(procs) == n:
            break

    return procs

# ─────────────────────────────────────────────
# Step 3 – Conflict Detection + Plan
# ─────────────────────────────────────────────
def build_rebalance_plan(
    core_stats: list[CoreStat],
    processes:  list[ProcessInfo],
) -> list[RebalanceAction]:
    """
    Flag PIDs crowding HOT cores while COLD cores are idle.
    Suggest moving them to COLD cores in round-robin.
    """
    hot_ids  = {c.core_id for c in core_stats if c.label == "HOT"}
    cold_ids = [c.core_id for c in core_stats if c.label == "COLD"]

    core_pid_map: dict[int, list] = {c.core_id: [] for c in core_stats}
    for proc in processes:
        for cid in proc.current_cores:
            if cid in core_pid_map:
                core_pid_map[cid].append(proc.pid)

    actions: list[RebalanceAction] = []
    cold_iter = iter(cold_ids * (len(processes) + 1))

    for proc in processes:
        on_hot  = [c for c in proc.current_cores if c in hot_ids]
        has_conflict = (
            len(on_hot) > 0 and
            len(cold_ids) > 0 and
            len(proc.current_cores) < len(core_stats)
        )

        if not has_conflict:
            continue

        try:
            to_cores = [next(cold_iter) for _ in proc.current_cores]
        except StopIteration:
            to_cores = cold_ids[:len(proc.current_cores)] or cold_ids[:1]

        actions.append(RebalanceAction(
            pid=proc.pid, name=proc.name,
            cpu_percent=proc.cpu_percent,
            from_cores=proc.current_cores,
            to_cores=to_cores,
        ))

    return actions

# ─────────────────────────────────────────────
# Step 4 – Execution
# ─────────────────────────────────────────────
def apply_action(action: RebalanceAction) -> RebalanceAction:
    """Apply taskset; mark as manual_only on failure."""
    cpu_list = cores_to_cpulist(action.to_cores)
    rc, out, err = run(["taskset", "-pc", cpu_list, str(action.pid)])
    if rc != 0:
        action.manual_only = True
        action.error_msg   = err or out or "Permission denied / kernel thread"
    return action

# ─────────────────────────────────────────────
# UI – Pretty Printers
# ─────────────────────────────────────────────
BANNER = f"""
{C.CYAN}{C.BOLD}╔══════════════════════════════════════════════════════════╗
║          TidyCPU  —  CPU Affinity Optimizer              ║
║          Linux x64  |  Requires root                     ║
╚══════════════════════════════════════════════════════════╝{C.RESET}"""

def usage_bar(pct: float, width: int = 20) -> str:
    filled = int(pct / 100 * width)
    bar    = "█" * filled + "░" * (width - filled)
    if   pct >= 80: color = C.RED
    elif pct >= 40: color = C.YELLOW
    else:           color = C.GREEN
    return f"{color}{bar}{C.RESET}"

def label_color(label: str) -> str:
    return {
        "HOT":  f"{C.RED}{C.BOLD}HOT {C.RESET}",
        "WARM": f"{C.YELLOW}WARM{C.RESET}",
        "COLD": f"{C.CYAN}COLD{C.RESET}",
    }.get(label, label)

def print_system_info(sysinfo: SystemInfo):
    """Print system hardware information."""
    print(f"\n{C.BOLD}{'─'*78}{C.RESET}")
    print(f"{C.BOLD}  SYSTEM INFORMATION{C.RESET}")
    print(f"{C.BOLD}{'─'*78}{C.RESET}")
    
    print(f"  {C.CYAN}CPU Model:{C.RESET}")
    print(f"    {sysinfo.cpu_model}")
    
    print(f"\n  {C.CYAN}Memory:{C.RESET}")
    print(f"    Total:     {C.GREEN}{sysinfo.total_memory}{C.RESET}")
    print(f"    Available: {C.GREEN}{sysinfo.available_memory}{C.RESET}")
    
    if sysinfo.cpu_freq_cur is not None:
        print(f"\n  {C.CYAN}CPU Frequency:{C.RESET}")
        if sysinfo.cpu_freq_min and sysinfo.cpu_freq_max:
            print(f"    Range:   {sysinfo.cpu_freq_min:.0f} MHz - {sysinfo.cpu_freq_max:.0f} MHz")
        print(f"    Current: {C.YELLOW}{sysinfo.cpu_freq_cur:.0f} MHz{C.RESET}")
    
    print(f"\n  {C.CYAN}Kernel Command Line:{C.RESET}")
    cmdline = sysinfo.kernel_cmdline
    if len(cmdline) > 72:
        words = cmdline.split()
        lines = []
        current_line = "    "
        for word in words:
            if len(current_line) + len(word) + 1 > 78:
                lines.append(current_line)
                current_line = "    " + word
            else:
                current_line += (" " if len(current_line) > 4 else "") + word
        if current_line.strip():
            lines.append(current_line)
        for line in lines:
            print(f"{C.DIM}{line}{C.RESET}")
    else:
        print(f"    {C.DIM}{cmdline}{C.RESET}")

# ── Table rendering helpers ────────────────────────────────────────────────────

# Column display widths (visible characters, excluding border/padding)
_COL_WIDTHS: dict[str, int] = {
    "Bar":     22,
    "Usage":   6,
    "Process": 15,
    "Parent":  15,
    "Core":    5,
}

# Default column order: left half outside→center, right half center→outside
_LEFT_ORDER  = ["Bar", "Usage", "Process", "Parent", "Core"]
_RIGHT_ORDER = ["Core", "Parent", "Process", "Usage", "Bar"]

# Dark-mode border color (dark gray)
_BD = "\033[90m"

def _hdr_color(col: str) -> str:
    return {"Bar": C.DIM, "Usage": C.RED, "Process": C.BLUE,
            "Parent": C.MAGENTA, "Core": C.CYAN}.get(col, "")

def _tbl_line(cols: list[str], lc: str, mc: str, rc: str) -> str:
    """Build a horizontal border line (top / separator / bottom)."""
    parts = [_BD + lc]
    for i, col in enumerate(cols):
        parts.append("─" * (_COL_WIDTHS[col] + 2))
        parts.append(mc if i < len(cols) - 1 else rc + C.RESET)
    return "".join(parts)

def _tbl_header(cols: list[str]) -> str:
    """Build the header row with colored column labels."""
    bd = _BD + "│" + C.RESET
    parts = [bd]
    for col in cols:
        w = _COL_WIDTHS[col]
        parts.append(f" {_hdr_color(col)}{C.BOLD}{col:^{w}}{C.RESET} " + bd)
    return "".join(parts)

def _render_cell(col: str, cs: Optional[CoreStat],
                 is_right: bool = False, hl: bool = False) -> str:
    """Render a single cell's content (no border chars)."""
    w = _COL_WIDTHS[col]
    pad = w + 2  # visible width including spaces
    if cs is None:
        return " " * pad
    bold = C.BOLD if hl else ""
    if col == "Bar":
        return f" {usage_bar(cs.usage, width=w)} "
    if col == "Usage":
        clr = C.RED if cs.usage >= 80 else (C.YELLOW if cs.usage >= 40 else C.GREEN)
        return f" {bold}{clr}{cs.usage:>5.1f}%{C.RESET} "
    if col == "Process":
        txt = (cs.top_proc or "—")[:w]
        clr = C.WHITE if hl else C.BLUE
        fmt = f">{w}" if is_right else f"<{w}"
        return f" {bold}{clr}{txt:{fmt}}{C.RESET} "
    if col == "Parent":
        txt = (cs.top_parent or "—")[:w]
        clr = C.WHITE if hl else C.MAGENTA
        fmt = f">{w}" if is_right else f"<{w}"
        return f" {bold}{clr}{txt:{fmt}}{C.RESET} "
    if col == "Core":
        txt = f"CPU{cs.core_id}"
        fmt = f">{w}" if is_right else f"<{w}"
        return f" {bold}{C.DIM}{txt:{fmt}}{C.RESET} "
    return " " * pad

def _tbl_row(cells: list[str]) -> str:
    """Wrap pre-rendered cells with border characters."""
    bd = _BD + "│" + C.RESET
    return bd + bd.join(cells) + bd

def _render_stacked_cell(col: str, proc_entry: Optional[tuple],
                         is_right: bool = False) -> str:
    """Render a cell for a stacked process sub-row (below the main core row)."""
    w = _COL_WIDTHS[col]
    pad = w + 2
    if proc_entry is None:
        return " " * pad
    thread_name, parent_name, cpu_pct = proc_entry
    if col == "Process":
        txt = thread_name[:w]
        fmt = f">{w}" if is_right else f"<{w}"
        return f" {C.DIM}{C.BLUE}{txt:{fmt}}{C.RESET} "
    if col == "Parent":
        txt = parent_name[:w]
        fmt = f">{w}" if is_right else f"<{w}"
        return f" {C.DIM}{C.MAGENTA}{txt:{fmt}}{C.RESET} "
    if col == "Usage":
        clr = C.RED if cpu_pct >= 80 else (C.YELLOW if cpu_pct >= 40 else C.GREEN)
        return f" {C.DIM}{clr}{cpu_pct:>5.1f}%{C.RESET} "
    # Bar and Core: empty for stacked rows
    return " " * pad

def print_topology(topology: dict[int, CPUTopology], core_stats: list[CoreStat],
                   ignore_cols: Optional[list[str]] = None,
                   specify_parents: Optional[list[str]] = None,
                   stack_procs: bool = False):
    """Print CPU topology table — two-column for HT servers, single-column otherwise."""
    ignored = {c.lower() for c in (ignore_cols or [])}
    vis_left  = [c for c in _LEFT_ORDER  if c.lower() not in ignored]
    vis_right = [c for c in _RIGHT_ORDER if c.lower() not in ignored]
    spec_set  = set(specify_parents or [])

    stats_map = {cs.core_id: cs for cs in core_stats}
    total_cores = len(topology)

    by_physical: dict = {}
    for logical_id, topo in topology.items():
        by_physical.setdefault(topo.physical_id, []).append(topo)

    total_physical = len(by_physical)
    total_physical_cores = sum(
        len(set(t.core_id for t in cores)) for cores in by_physical.values()
    )
    ht_enabled = total_cores > total_physical_cores

    print(f"\n{C.BOLD}  CPU TOPOLOGY{C.RESET}\n")

    if ht_enabled:
        all_cols = vis_left + vis_right
        left_count = (total_cores + 1) // 2

        print("  " + _tbl_line(all_cols, "┌", "┬", "┐"))
        print("  " + _tbl_header(all_cols))
        print("  " + _tbl_line(all_cols, "├", "┼", "┤"))

        for i in range(left_count):
            cs_l = stats_map.get(i)
            cs_r = stats_map.get(i + left_count)
            hl_l = bool(spec_set and cs_l and cs_l.top_parent in spec_set)
            hl_r = bool(spec_set and cs_r and cs_r.top_parent in spec_set)
            cells = (
                [_render_cell(c, cs_l, is_right=False, hl=hl_l) for c in vis_left] +
                [_render_cell(c, cs_r, is_right=True,  hl=hl_r) for c in vis_right]
            )
            print("  " + _tbl_row(cells))

            if stack_procs:
                procs_l = cs_l.all_procs if cs_l else []
                procs_r = cs_r.all_procs if cs_r else []
                for j in range(max(len(procs_l), len(procs_r))):
                    entry_l = procs_l[j] if j < len(procs_l) else None
                    entry_r = procs_r[j] if j < len(procs_r) else None
                    stacked = (
                        [_render_stacked_cell(c, entry_l, is_right=False) for c in vis_left] +
                        [_render_stacked_cell(c, entry_r, is_right=True)  for c in vis_right]
                    )
                    print("  " + _tbl_row(stacked))

        print("  " + _tbl_line(all_cols, "└", "┴", "┘"))
    else:
        print("  " + _tbl_line(vis_left, "┌", "┬", "┐"))
        print("  " + _tbl_header(vis_left))
        print("  " + _tbl_line(vis_left, "├", "┼", "┤"))

        for i in range(total_cores):
            cs = stats_map.get(i)
            hl = bool(spec_set and cs and cs.top_parent in spec_set)
            cells = [_render_cell(c, cs, is_right=False, hl=hl) for c in vis_left]
            print("  " + _tbl_row(cells))

            if stack_procs and cs:
                for entry in cs.all_procs:
                    stacked = [_render_stacked_cell(c, entry, is_right=False) for c in vis_left]
                    print("  " + _tbl_row(stacked))

        print("  " + _tbl_line(vis_left, "└", "┴", "┘"))

    hot  = sum(1 for c in core_stats if c.label == "HOT")
    warm = sum(1 for c in core_stats if c.label == "WARM")
    cold = sum(1 for c in core_stats if c.label == "COLD")

    if spec_set:
        print(f"\n  {C.BOLD}Focused on:{C.RESET} {', '.join(sorted(spec_set))}")

    print(f"\n  Summary: {C.RED}●{C.RESET} {hot} Hot  "
          f"{C.YELLOW}●{C.RESET} {warm} Warm  "
          f"{C.CYAN}●{C.RESET} {cold} Cold  {C.DIM}|{C.RESET}  "
          f"{total_physical} physical CPU(s), "
          f"{total_physical_cores} physical core(s), "
          f"{total_cores} logical core(s)")

    ht_status = f"{C.GREEN}Enabled{C.RESET}" if ht_enabled else f"{C.DIM}Disabled{C.RESET}"
    print(f"  Hyperthreading: {ht_status}")

def clear_screen():
    """Clear terminal screen."""
    os.system('clear' if os.name != 'nt' else 'cls')

def _fetch_process_info(pid: int, num_cores: int) -> Optional[ProcessInfo]:
    """Fetch a single ProcessInfo for the given PID. Returns None if unavailable."""
    rc, out, _ = run(["ps", "-p", str(pid), "-o", "pid,pcpu,comm", "--no-headers"])
    if rc != 0 or not out:
        return None
    parts = out.strip().split(None, 2)
    if len(parts) < 3:
        return None
    try:
        pid_val = int(parts[0])
        cpu     = float(parts[1])
        name    = parts[2].strip()[:24]
    except ValueError:
        return None

    rc2, out2, _ = run(["taskset", "-p", str(pid_val)])
    affinity_mask = "N/A"
    cur_cores: list[int] = []
    if rc2 == 0:
        m = re.search(r"current affinity mask:\s*([0-9a-fA-F,]+)", out2)
        if m:
            affinity_mask = m.group(1)
            cur_cores = mask_to_cores(affinity_mask, num_cores)

    threads = get_threads_for_pid(pid_val, num_cores)
    return ProcessInfo(
        pid=pid_val, name=name, cpu_percent=cpu,
        current_cores=cur_cores, affinity_mask=affinity_mask,
        threads=threads
    )

def live_monitor(duration_sec: int = 5, interval_ms: int = 3000, filter_pids: Optional[list[int]] = None, show_cpu_freq: bool = False, export_html: Optional[str] = None, export_text: Optional[str] = None, export_excel: Optional[str] = None, sysinfo: Optional[SystemInfo] = None, topology_data: Optional[dict] = None, ignore_cols: Optional[list[str]] = None, specify_parents: Optional[list[str]] = None, stack_procs: bool = False, min_usage: float = 0.0, ignore_procs: Optional[list[str]] = None):
    """
    Live monitoring mode - refresh stats every interval_ms for duration_sec iterations.
    CPU is sampled for SAMPLE_MS (fast), then the screen is shown for the remaining
    display time so the user can actually read it before the next refresh.
    If filter_pids is set, show threads for each of those PIDs.
    Collects snapshots for export if export options are provided.
    """
    SAMPLE_MS = 500   # CPU sampling window — short, accurate
    display_ms = max(interval_ms - SAMPLE_MS, 1000)  # how long screen stays visible

    topology = topology_data if topology_data else get_cpu_topology()
    num_cores = len(topology)

    iterations = duration_sec
    snapshots = [] if (export_html or export_text or export_excel) else None

    for i in range(iterations):
        # ── 1. Sample CPU (happens invisibly, screen still showing previous frame) ──
        core_stats = get_core_usage(sample_ms=SAMPLE_MS, topology=topology,
                                    stack_procs=stack_procs, min_usage=min_usage,
                                    ignore_procs=ignore_procs)

        # ── 2. Collect process info ───────────────────────────────────────────────
        processes: list[ProcessInfo] = []
        if filter_pids:
            for pid in filter_pids:
                info = _fetch_process_info(pid, num_cores)
                if info is not None:
                    processes.append(info)

        # ── 3. Render (clear → print → user reads) ────────────────────────────────
        clear_screen()
        print(BANNER)

        if i == 0 and sysinfo:
            print_system_info(sysinfo)

        refresh_label = f"{interval_ms // 1000}s"
        print(f"\n  {C.CYAN}Live Monitor Mode{C.RESET} — {C.DIM}Refresh: {refresh_label}  "
              f"Iteration: {i+1}/{iterations}{C.RESET}")

        print_topology(topology, core_stats,
                       ignore_cols=ignore_cols,
                       specify_parents=specify_parents,
                       stack_procs=stack_procs)

        if filter_pids:
            label = " | ".join(str(p) for p in filter_pids)
            print(f"\n  {C.YELLOW}Filtering: {label}{C.RESET}")
            for info in processes:
                print_process_details_live(info)
            not_found = [p for p in filter_pids if not any(pr.pid == p for pr in processes)]
            for pid in not_found:
                print(f"  {C.RED}✘ PID {pid} not found or inaccessible{C.RESET}")

        if snapshots is not None:
            snapshot = Snapshot(
                timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                iteration=i + 1,
                core_stats=core_stats,
                processes=processes,
            )
            snapshots.append(snapshot)

        print(f"\n  {C.DIM}Press Ctrl+C to exit{C.RESET}")

        # ── 4. Sleep the display window so the user can read the screen ───────────
        if i < iterations - 1:
            time.sleep(display_ms / 1000)
    
    print(f"\n  {C.GREEN}Live monitoring complete.{C.RESET}\n")
    
    if snapshots and sysinfo:
        if export_html:
            try:
                filename = export_to_html(sysinfo, topology, core_stats, [], export_html,
                                          snapshots=snapshots, ignore_cols=ignore_cols)
                print(f"  {C.GREEN}✔{C.RESET}  HTML report with {len(snapshots)} snapshots exported to: {C.CYAN}{filename}{C.RESET}\n")
            except Exception as e:
                print(f"  {C.RED}✘{C.RESET}  HTML export failed: {e}\n")

        if export_text:
            try:
                filename = export_to_text(sysinfo, topology, core_stats, [], export_text,
                                          snapshots=snapshots, ignore_cols=ignore_cols)
                print(f"  {C.GREEN}✔{C.RESET}  Text report with {len(snapshots)} snapshots exported to: {C.CYAN}{filename}{C.RESET}\n")
            except Exception as e:
                print(f"  {C.RED}✘{C.RESET}  Text export failed: {e}\n")

        if export_excel:
            try:
                filename = export_to_excel(sysinfo, topology, core_stats, [], export_excel,
                                           snapshots=snapshots, ignore_cols=ignore_cols,
                                           stack_procs=stack_procs)
                print(f"  {C.GREEN}✔{C.RESET}  Excel report with {len(snapshots)} snapshots exported to: {C.CYAN}{filename}{C.RESET}\n")
            except Exception as e:
                print(f"  {C.RED}✘{C.RESET}  Excel export failed: {e}\n")

def print_process_details_live(proc: ProcessInfo):
    """Print a single process's thread details during live monitoring (no affinity mask)."""
    print(f"\n{C.BOLD}{'─'*62}{C.RESET}")
    print(f"{C.BOLD}  PROCESS: {proc.name}  (PID {proc.pid})  —  {C.YELLOW}{proc.cpu_percent:.1f}% CPU{C.RESET}")
    print(f"{C.BOLD}{'─'*62}{C.RESET}")

    if proc.threads:
        print(f"  {'TID':>7}  {'Name':<24}  {'CPU%':>6}  Cores")
        print(f"  {'───':>7}  {'────────────────────────':<24}  {'─────':>6}  ─────")
        for t in proc.threads[:15]:
            t_cores_str = ",".join(map(str, t.current_cores)) if t.current_cores else "all"
            print(
                f"  {t.tid:>7}  {t.name:<24}  "
                f"{C.YELLOW}{t.cpu_percent:>5.1f}%{C.RESET}  {t_cores_str}"
            )
        if len(proc.threads) > 15:
            print(f"  {C.DIM}... {len(proc.threads) - 15} more threads{C.RESET}")

def print_rebalance_plan(actions: list[RebalanceAction]):
    print(f"\n{C.BOLD}{'─'*62}{C.RESET}")
    print(f"{C.BOLD}  REBALANCING PLAN{C.RESET}")
    print(f"{C.BOLD}{'─'*62}{C.RESET}")
    if not actions:
        print(f"  {C.GREEN}✔  No conflicts detected. CPU load looks balanced!{C.RESET}")
        return

    print(f"  {'PID':>7}  {'Process':<20}  {'CPU%':>5}  {'From → To'}")
    print(f"  {'───────':>7}  {'────────────────────':20}  {'─────':>5}  {'──────────────────────'}")
    for a in actions:
        from_s = ",".join(map(str, a.from_cores))
        to_s   = ",".join(map(str, a.to_cores))
        arrow  = (
            f"{C.RED}Core [{from_s}]{C.RESET}"
            f" {C.DIM}→{C.RESET} "
            f"{C.GREEN}Core [{to_s}]{C.RESET}"
        )
        print(
            f"  {C.MAGENTA}{a.pid:>7}{C.RESET}  "
            f"{a.name:<20}  "
            f"{C.YELLOW}{a.cpu_percent:>4.1f}%{C.RESET}  "
            f"{arrow}"
        )
    print(f"\n  {len(actions)} change(s) proposed.")

def print_results(actions: list[RebalanceAction]):
    print(f"\n{C.BOLD}{'─'*62}{C.RESET}")
    print(f"{C.BOLD}  EXECUTION RESULTS{C.RESET}")
    print(f"{C.BOLD}{'─'*62}{C.RESET}")
    ok      = [a for a in actions if not a.manual_only]
    manual  = [a for a in actions if a.manual_only]

    for a in ok:
        to_s = ",".join(map(str, a.to_cores))
        print(f"  {C.GREEN}✔{C.RESET}  PID {C.MAGENTA}{a.pid}{C.RESET} ({a.name})  →  pinned to core(s) [{to_s}]")

    if manual:
        print(f"\n  {C.YELLOW}⚠  Manual Suggestions (could not auto-apply):{C.RESET}")
        for a in manual:
            to_s = ",".join(map(str, a.to_cores))
            print(
                f"  {C.YELLOW}↪{C.RESET}  PID {C.MAGENTA}{a.pid}{C.RESET} ({a.name})"
                f"  — run manually:\n"
                f"       {C.DIM}sudo taskset -pc {to_s} {a.pid}{C.RESET}\n"
                f"       {C.RED}Reason: {a.error_msg}{C.RESET}"
            )

# ─────────────────────────────────────────────
# Export Functions
# ─────────────────────────────────────────────
def export_to_html(
    sysinfo: SystemInfo,
    topology: dict[int, CPUTopology],
    core_stats: list[CoreStat],
    processes: list[ProcessInfo],
    filename: str = "tidycpu_report.html",
    snapshots: Optional[list[Snapshot]] = None,
    ignore_cols: Optional[list[str]] = None,
):
    """Export current state to HTML file with styling."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>TidyCPU Report - {timestamp}</title>
    <style>
        body {{
            font-family: 'Consolas', 'Monaco', monospace;
            background: #1e1e1e;
            color: #d4d4d4;
            padding: 20px;
            line-height: 1.6;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        h1 {{ color: #4fc3f7; border-bottom: 2px solid #4fc3f7; padding-bottom: 10px; }}
        h2 {{ color: #81c784; margin-top: 30px; }}
        h3 {{ color: #ffb74d; margin-top: 20px; }}
        .section {{
            background: #252526; padding: 20px; margin: 20px 0;
            border-radius: 5px; border-left: 4px solid #4fc3f7;
        }}
        .snapshot {{
            background: #2d2d30; padding: 15px; margin: 15px 0;
            border-radius: 5px; border-left: 4px solid #ffb74d;
        }}
        .info-grid {{ display: grid; grid-template-columns: 200px 1fr; gap: 10px; margin: 10px 0; }}
        .label {{ color: #4fc3f7; font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 14px; }}
        th {{ background: #2d2d30; color: #4fc3f7; padding: 10px; text-align: left; border-bottom: 2px solid #4fc3f7; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #3e3e42; }}
        tr:hover {{ background: #2d2d30; }}
        .bar {{ background: #3e3e42; height: 20px; border-radius: 3px; overflow: hidden; }}
        .bar-fill {{ height: 100%; }}
        .bar-fill.hot {{ background: #f44336; }}
        .bar-fill.warm {{ background: #ff9800; }}
        .bar-fill.cold {{ background: #4caf50; }}
        .status {{ padding: 3px 8px; border-radius: 3px; font-weight: bold; font-size: 12px; }}
        .status.hot {{ background: #f44336; color: white; }}
        .status.warm {{ background: #ff9800; color: white; }}
        .status.cold {{ background: #4caf50; color: white; }}
        .summary {{ display: flex; gap: 20px; margin: 15px 0; flex-wrap: wrap; }}
        .summary-item {{ background: #2d2d30; padding: 10px 15px; border-radius: 5px; }}
        .dot {{ display: inline-block; width: 10px; height: 10px; border-radius: 50%; margin-right: 5px; }}
        .dot.hot {{ background: #f44336; }}
        .dot.warm {{ background: #ff9800; }}
        .dot.cold {{ background: #4caf50; }}
        .timestamp {{ color: #858585; font-size: 12px; }}
        .two-column {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
        .iteration-badge {{
            display: inline-block; background: #ffb74d; color: #1e1e1e;
            padding: 5px 10px; border-radius: 5px; font-weight: bold; margin-left: 10px;
        }}
        .process-name {{ color: #ce9178; font-style: italic; }}
        .parent-name  {{ color: #c586c0; font-style: italic; }}
        .tabs {{ display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 16px; }}
        .tab-btn {{
            background: #2d2d30; color: #9cdcfe; border: 1px solid #3e3e42;
            padding: 6px 14px; border-radius: 4px; cursor: pointer; font-family: inherit;
            font-size: 13px; transition: background 0.15s;
        }}
        .tab-btn:hover {{ background: #3e3e42; }}
        .tab-btn.active {{ background: #ffb74d; color: #1e1e1e; border-color: #ffb74d; font-weight: bold; }}
        .tab-panel {{ display: none; }}
        .tab-panel.active {{ display: block; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>TidyCPU — CPU Affinity Optimization Report</h1>
        <p class="timestamp">Generated: {timestamp}</p>
"""
    
    # System Information
    html += """
        <div class="section">
            <h2>System Information</h2>
            <div class="info-grid">
                <div class="label">CPU Model:</div>
                <div>{}</div>
                <div class="label">Total Memory:</div>
                <div>{}</div>
                <div class="label">Available Memory:</div>
                <div>{}</div>
""".format(sysinfo.cpu_model, sysinfo.total_memory, sysinfo.available_memory)
    
    if sysinfo.cpu_freq_cur:
        html += f"""
                <div class="label">CPU Frequency:</div>
                <div>{sysinfo.cpu_freq_min:.0f} MHz - {sysinfo.cpu_freq_max:.0f} MHz (Current: {sysinfo.cpu_freq_cur:.0f} MHz)</div>
"""
    
    html += f"""
                <div class="label">Kernel Cmdline:</div>
                <div style="word-break: break-all;">{sysinfo.kernel_cmdline}</div>
            </div>
        </div>
"""
    
    def render_topology_section(core_stats_data, title_suffix=""):
        ignored = {c.lower() for c in (ignore_cols or [])}
        vis_left  = [c for c in _LEFT_ORDER  if c.lower() not in ignored]
        vis_right = [c for c in _RIGHT_ORDER if c.lower() not in ignored]

        stats_map = {cs.core_id: cs for cs in core_stats_data}
        total_cores = len(stats_map)
        left_count = (total_cores + 1) // 2

        hot_count  = sum(1 for c in core_stats_data if c.label == "HOT")
        warm_count = sum(1 for c in core_stats_data if c.label == "WARM")
        cold_count = sum(1 for c in core_stats_data if c.label == "COLD")

        by_physical = {}
        for _lid, topo in topology.items():
            by_physical.setdefault(topo.physical_id, []).append(topo)

        total_physical = len(by_physical)
        total_physical_cores = sum(len(set(t.core_id for t in v)) for v in by_physical.values())
        ht_enabled = total_cores > total_physical_cores

        s = f"""
            <h3>CPU Topology{title_suffix}</h3>
            <div class="summary">
                <div class="summary-item"><span class="dot hot"></span>{hot_count} Hot</div>
                <div class="summary-item"><span class="dot warm"></span>{warm_count} Warm</div>
                <div class="summary-item"><span class="dot cold"></span>{cold_count} Cold</div>
                <div class="summary-item">{total_physical} Physical CPU(s)</div>
                <div class="summary-item">{total_physical_cores} Physical Cores</div>
                <div class="summary-item">{total_cores} Logical Cores</div>
                <div class="summary-item">HT: {'Enabled' if ht_enabled else 'Disabled'}</div>
            </div>
"""
        def _cell(col, cs):
            lc = cs.label.lower()
            if col == "Bar":
                return (f"<td><div class=\"bar\"><div class=\"bar-fill {lc}\""
                        f" style=\"width:{cs.usage:.1f}%\"></div></div></td>")
            if col == "Usage":
                return f"<td>{cs.usage:.1f}%</td>"
            if col == "Process":
                return f"<td><span class=\"process-name\">{cs.top_proc or '&#8212;'}</span></td>"
            if col == "Parent":
                return f"<td><span class=\"parent-name\">{cs.top_parent or '&#8212;'}</span></td>"
            if col == "Core":
                return f"<td><code>CPU{cs.core_id}</code></td>"
            return "<td></td>"

        def _row(cs, cols):
            return "<tr>" + "".join(_cell(c, cs) for c in cols) + "</tr>\n"

        def _hdr(cols):
            return ("<thead><tr>"
                    + "".join(f"<th>{c}</th>" for c in cols)
                    + "</tr></thead>")

        if ht_enabled:
            s += '            <div class="two-column">\n'
            s += f"                <table>{_hdr(vis_left)}<tbody>\n"
            for i in range(left_count):
                if i in stats_map:
                    s += "                    " + _row(stats_map[i], vis_left)
            s += "                </tbody></table>\n"
            s += f"                <table>{_hdr(vis_right)}<tbody>\n"
            for i in range(left_count, total_cores):
                if i in stats_map:
                    s += "                    " + _row(stats_map[i], vis_right)
            s += "                </tbody></table>\n"
            s += "            </div>\n"
        else:
            s += f"            <table>{_hdr(vis_left)}<tbody>\n"
            for i in range(total_cores):
                if i in stats_map:
                    s += "                " + _row(stats_map[i], vis_left)
            s += "            </tbody></table>\n"

        return s
    
    if snapshots:
        html += f"""
        <div class="section">
            <h2>Live Monitoring Results <span class="iteration-badge">{len(snapshots)} Snapshots</span></h2>
            <div class="tabs">
"""
        for snap in snapshots:
            active = 'active' if snap.iteration == 1 else ''
            ts = snap.timestamp.split()[1]
            html += f'                <button class="tab-btn {active}" onclick="showTab({snap.iteration})" id="btn-{snap.iteration}">#{snap.iteration} &nbsp;<span style="color:#858585;font-size:11px">{ts}</span></button>\n'
        html += "            </div>\n"
        for snap in snapshots:
            active = 'active' if snap.iteration == 1 else ''
            html += f"""
            <div class="tab-panel {active}" id="panel-{snap.iteration}">
                {render_topology_section(snap.core_stats)}
            </div>
"""
        html += """
<script>
            function showTab(n) {{
                document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
                document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
                document.getElementById('panel-' + n).classList.add('active');
                document.getElementById('btn-' + n).classList.add('active');
            }}
            </script>
        </div>
"""
    else:
        html += f"""
        <div class="section">
            {render_topology_section(core_stats)}
        </div>
"""
    
    html += """
    </div>
</body>
</html>
"""
    
    with open(filename, 'w') as f:
        f.write(html)
    
    return filename

def export_to_text(
    sysinfo: SystemInfo,
    topology: dict[int, CPUTopology],
    core_stats: list[CoreStat],
    processes: list[ProcessInfo],
    filename: str = "tidycpu_report.txt",
    snapshots: Optional[list[Snapshot]] = None,
    ignore_cols: Optional[list[str]] = None,
):
    """Export current state to text file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    output = []
    output.append("=" * 78)
    output.append("TidyCPU — CPU Affinity Optimization Report")
    output.append(f"Generated: {timestamp}")
    output.append("=" * 78)
    output.append("")
    
    output.append("SYSTEM INFORMATION")
    output.append("-" * 78)
    output.append(f"CPU Model:        {sysinfo.cpu_model}")
    output.append(f"Total Memory:     {sysinfo.total_memory}")
    output.append(f"Available Memory: {sysinfo.available_memory}")
    
    if sysinfo.cpu_freq_cur:
        output.append(f"CPU Frequency:    {sysinfo.cpu_freq_min:.0f} MHz - {sysinfo.cpu_freq_max:.0f} MHz")
        output.append(f"                  (Current: {sysinfo.cpu_freq_cur:.0f} MHz)")
    
    output.append(f"Kernel Cmdline:   {sysinfo.kernel_cmdline}")
    output.append("")
    
    def render_topology(core_stats_data, title_suffix=""):
        ignored = {c.lower() for c in (ignore_cols or [])}
        vis_left  = [c for c in _LEFT_ORDER  if c.lower() not in ignored]
        vis_right = [c for c in _RIGHT_ORDER if c.lower() not in ignored]

        stats_map = {cs.core_id: cs for cs in core_stats_data}
        total_cores = len(stats_map)
        left_count = (total_cores + 1) // 2

        hot_count  = sum(1 for c in core_stats_data if c.label == "HOT")
        warm_count = sum(1 for c in core_stats_data if c.label == "WARM")
        cold_count = sum(1 for c in core_stats_data if c.label == "COLD")

        by_physical = {}
        for _lid, topo in topology.items():
            by_physical.setdefault(topo.physical_id, []).append(topo)

        total_physical = len(by_physical)
        total_physical_cores = sum(len(set(t.core_id for t in v)) for v in by_physical.values())
        ht_enabled = total_cores > total_physical_cores

        W = _COL_WIDTHS  # shared with terminal renderer

        def _bar(pct):
            f = int(pct / 100 * W["Bar"])
            return '█' * f + '░' * (W["Bar"] - f)

        def _cell(col, cs, is_right=False):
            if cs is None:
                return " " * W[col]
            w = W[col]
            if col == "Bar":
                return _bar(cs.usage)
            if col == "Usage":
                return f"{cs.usage:>{w-1}.1f}%"
            if col == "Process":
                t = (cs.top_proc or "—")[:w]
                return f"{t:>{w}}" if is_right else f"{t:<{w}}"
            if col == "Parent":
                t = (cs.top_parent or "—")[:w]
                return f"{t:>{w}}" if is_right else f"{t:<{w}}"
            if col == "Core":
                t = f"CPU{cs.core_id}"
                return f"{t:>{w}}" if is_right else f"{t:<{w}}"
            return " " * w

        def _row(cs, cols, is_right=False):
            return "  ".join(_cell(c, cs, is_right) for c in cols)

        def _hdr(cols, is_right=False):
            return "  ".join(
                f"{c:>{W[c]}}" if (is_right and c not in ("Bar", "Usage")) else f"{c:<{W[c]}}"
                for c in cols
            )

        def _sep(cols):
            return "  ".join("─" * W[c] for c in cols)

        lines = [f"CPU TOPOLOGY{title_suffix}"]

        if ht_enabled:
            lines.append(_hdr(vis_left) + "  " + _hdr(vis_right, is_right=True))
            lines.append(_sep(vis_left)  + "  " + _sep(vis_right))
            for i in range(left_count):
                left  = _row(stats_map.get(i), vis_left)
                right = _row(stats_map.get(i + left_count), vis_right, is_right=True)
                lines.append(left + ("  " + right if stats_map.get(i + left_count) else ""))
        else:
            lines.append(_hdr(vis_left))
            lines.append(_sep(vis_left))
            for i in range(total_cores):
                lines.append(_row(stats_map.get(i), vis_left))

        lines.append("")
        lines.append(f"Summary: ● {hot_count} Hot  ● {warm_count} Warm  ● {cold_count} Cold  |  "
                     f"{total_physical} physical CPU(s), {total_physical_cores} physical core(s), "
                     f"{total_cores} logical core(s)")
        lines.append(f"Hyperthreading: {'Enabled' if ht_enabled else 'Disabled'}")
        return lines
    
    if snapshots:
        output.append(f"LIVE MONITORING RESULTS ({len(snapshots)} snapshots)")
        output.append("=" * 78)
        output.append("")
        
        for snap in snapshots:
            output.append(f"[Iteration {snap.iteration}] @ {snap.timestamp}")
            output.append("-" * 78)
            output.extend(render_topology(snap.core_stats, f" - Iteration {snap.iteration}"))
            output.append("")
            output.append("=" * 78)
            output.append("")
    else:
        output.extend(render_topology(core_stats))
        output.append("")
    
    output.append("=" * 78)
    
    with open(filename, 'w') as f:
        f.write('\n'.join(output))
    
    return filename

def export_to_excel(
    sysinfo: SystemInfo,
    topology: dict[int, CPUTopology],
    core_stats: list[CoreStat],
    processes: list[ProcessInfo],
    filename: str = "tidycpu_report.xlsx",
    snapshots: Optional[list[Snapshot]] = None,
    ignore_cols: Optional[list[str]] = None,
    stack_procs: bool = False,
):
    """Export report to an Excel (.xlsx) file using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import DataBarRule
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel export. "
            "Install it with:  pip install openpyxl"
        )

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Colour helpers ───────────────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _font(hex_color: str = "000000", bold: bool = False,
              italic: bool = False, size: int = 11) -> Font:
        return Font(color=hex_color, bold=bold, italic=italic, size=size)

    def _border() -> Border:
        side = Side(style="thin", color="CCCCCC")
        return Border(left=side, right=side, top=side, bottom=side)

    def _center() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    FILL_HEADER   = _fill("1E3A5F")
    FILL_ALT_ROW  = _fill("F2F2F2")
    FILL_HOT      = _fill("F44336")
    FILL_WARM     = _fill("FF9800")
    FILL_COLD     = _fill("4CAF50")
    FILL_STACKED  = _fill("FAFAFA")   # very light gray for stacked sub-rows
    FONT_HDR      = _font("FFFFFF", bold=True)
    FONT_TITLE    = _font("1E3A5F", bold=True, size=14)
    FONT_WHITE    = _font("FFFFFF", bold=True)
    FONT_DIM      = _font("666666", italic=True)
    FONT_STACKED  = _font("999999", italic=True)   # dimmed italic for stacked rows

    STATUS_STYLE = {
        "HOT":  (FILL_HOT,  FONT_WHITE),
        "WARM": (FILL_WARM, FONT_WHITE),
        "COLD": (FILL_COLD, FONT_WHITE),
    }

    def _auto_width(ws):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8,
            )
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 55)

    def _set_header_row(ws, row: int, headers: list[str]):
        ws.row_dimensions[row].height = 20
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font      = FONT_HDR
            cell.fill      = FILL_HEADER
            cell.border    = _border()
            cell.alignment = _center()

    # ── Detect HT ────────────────────────────────────────────────────────────
    by_physical: dict = {}
    for _lid, topo in topology.items():
        by_physical.setdefault(topo.physical_id, []).append(topo)

    total_physical       = len(by_physical)
    total_physical_cores = sum(len(set(t.core_id for t in v)) for v in by_physical.values())
    total_logical        = len(topology)
    ht_enabled           = total_logical > total_physical_cores

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: System Information ───────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "System Info"
    ws_info.column_dimensions["A"].width = 24
    ws_info.column_dimensions["B"].width = 62

    ws_info["A1"] = "TidyCPU — CPU Affinity Optimization Report"
    ws_info["A1"].font = FONT_TITLE
    ws_info.merge_cells("A1:B1")

    ws_info["A2"] = f"Generated: {timestamp}"
    ws_info["A2"].font = FONT_DIM
    ws_info.merge_cells("A2:B2")

    info_rows = [
        ("CPU Model",         sysinfo.cpu_model),
        ("Total Memory",      sysinfo.total_memory),
        ("Available Memory",  sysinfo.available_memory),
        ("Kernel Cmdline",    sysinfo.kernel_cmdline),
        ("Physical CPU(s)",   total_physical),
        ("Physical Core(s)",  total_physical_cores),
        ("Logical Core(s)",   total_logical),
        ("Hyperthreading",    "Enabled" if ht_enabled else "Disabled"),
    ]
    if sysinfo.cpu_freq_cur is not None:
        if sysinfo.cpu_freq_min and sysinfo.cpu_freq_max:
            info_rows.append(("CPU Freq Range",
                               f"{sysinfo.cpu_freq_min:.0f} – {sysinfo.cpu_freq_max:.0f} MHz"))
        info_rows.append(("CPU Freq Current", f"{sysinfo.cpu_freq_cur:.0f} MHz"))

    _set_header_row(ws_info, 4, ["Property", "Value"])
    for r_idx, (prop, val) in enumerate(info_rows, start=5):
        cell_a = ws_info.cell(row=r_idx, column=1, value=prop)
        cell_b = ws_info.cell(row=r_idx, column=2, value=str(val))
        cell_a.font = _font(bold=True)
        if r_idx % 2 == 0:
            cell_a.fill = FILL_ALT_ROW
            cell_b.fill = FILL_ALT_ROW
        cell_a.border = _border()
        cell_b.border = _border()
        cell_b.alignment = Alignment(wrap_text=True, vertical="center")

    # ── Helper: write one topology sheet ─────────────────────────────────────
    def _write_topology_sheet(ws, core_stats_data: list[CoreStat]):
        ignored = {c.lower() for c in (ignore_cols or [])}

        # Build ordered list of (excel_header, value_fn, col_type)
        # "Bar" controls the DataBar on the Usage column — not a separate column.
        all_col_defs = [
            ("Core",      lambda cs: f"CPU{cs.core_id}",          "core"),
            ("Usage (%)", lambda cs: round(cs.usage, 1),           "usage"),
            ("Process",   lambda cs: cs.top_proc    or "—",        "process"),
            ("Parent",    lambda cs: cs.top_parent  or "—",        "parent"),
        ]
        col_defs = [(hdr, fn, kind) for hdr, fn, kind in all_col_defs
                    if kind not in ignored]
        show_databar = "bar" not in ignored and any(k == "usage" for _, _, k in col_defs)

        headers = [h for h, _, _ in col_defs]
        if stack_procs:
            headers = ["Core", "Usage (%)", "Process", "Parent", "Type"]
        _set_header_row(ws, 1, headers)

        row_idx = 2
        for cs in sorted(core_stats_data, key=lambda x: x.core_id):
            # ── Main core row ─────────────────────────────────────────────────
            for col_idx, (_, fn, kind) in enumerate(col_defs, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=fn(cs))
                cell.border = _border()
                if kind == "core":
                    cell.alignment = _center()
                    if cs.label in STATUS_STYLE:
                        cell.fill, cell.font = STATUS_STYLE[cs.label]
                elif kind == "usage":
                    cell.alignment = _center()
                if row_idx % 2 == 0 and kind != "core":
                    cell.fill = FILL_ALT_ROW

            if stack_procs:
                # Write "Top" label in the Type column (one past col_defs)
                type_col = len(col_defs) + 1
                type_cell = ws.cell(row=row_idx, column=type_col, value="top")
                type_cell.font = _font("1E3A5F", bold=True)
                type_cell.border = _border()
                type_cell.alignment = _center()
            row_idx += 1

            # ── Stacked sub-rows (one per process in all_procs) ───────────────
            if stack_procs:
                col_map = {kind: idx for idx, (_, _, kind) in enumerate(col_defs, start=1)}
                type_col = len(col_defs) + 1
                for thread_name, parent_name, cpu_pct in cs.all_procs:
                    # Core cell: empty (already identified by the main row above)
                    if "core" in col_map:
                        c = ws.cell(row=row_idx, column=col_map["core"], value="")
                        c.border = _border()
                        c.alignment = _center()
                    # Usage: per-thread CPU%
                    if "usage" in col_map:
                        c = ws.cell(row=row_idx, column=col_map["usage"],
                                    value=round(cpu_pct, 1))
                        c.font = FONT_STACKED
                        c.fill = FILL_STACKED
                        c.border = _border()
                        c.alignment = _center()
                    # Process
                    if "process" in col_map:
                        c = ws.cell(row=row_idx, column=col_map["process"],
                                    value=thread_name or "—")
                        c.font = FONT_STACKED
                        c.fill = FILL_STACKED
                        c.border = _border()
                    # Parent
                    if "parent" in col_map:
                        c = ws.cell(row=row_idx, column=col_map["parent"],
                                    value=parent_name or "—")
                        c.font = FONT_STACKED
                        c.fill = FILL_STACKED
                        c.border = _border()
                    # Type label
                    tc = ws.cell(row=row_idx, column=type_col, value="stacked")
                    tc.font = FONT_STACKED
                    tc.fill = FILL_STACKED
                    tc.border = _border()
                    tc.alignment = _center()
                    row_idx += 1

        # Native DataBar on the Usage column (only when Bar is not hidden)
        if show_databar:
            usage_col_idx = next(i for i, (_, _, k) in enumerate(col_defs, 1) if k == "usage")
            usage_letter  = get_column_letter(usage_col_idx)
            last_row = row_idx - 1
            if last_row >= 2:
                ws.conditional_formatting.add(
                    f"{usage_letter}2:{usage_letter}{last_row}",
                    DataBarRule(start_type="num", start_value=0,
                                end_type="num", end_value=100,
                                color="4472C4"),
                )
        _auto_width(ws)

    # ── Write snapshot sheets (or single topology sheet) ─────────────────────
    if snapshots:
        for snap in snapshots:
            sheet_name = f"Snap {snap.iteration} ({snap.timestamp.split()[1].replace(':', '-')})"
            ws_snap = wb.create_sheet(title=sheet_name[:31])   # Excel: max 31 chars
            _write_topology_sheet(ws_snap, snap.core_stats)
    else:
        ws_topo = wb.create_sheet(title="CPU Topology")
        _write_topology_sheet(ws_topo, core_stats)

    # ── Sheet: Processes ──────────────────────────────────────────────────────
    if processes:
        ws_proc = wb.create_sheet(title="Processes")
        proc_headers = ["PID", "Name", "CPU (%)", "Cores", "Affinity Mask"]
        _set_header_row(ws_proc, 1, proc_headers)

        for row_idx, proc in enumerate(processes, start=2):
            cores_str = (", ".join(map(str, proc.current_cores))
                         if proc.current_cores else "all")
            row_data = [proc.pid, proc.name,
                        round(proc.cpu_percent, 1), cores_str, proc.affinity_mask]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_proc.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _border()
                if row_idx % 2 == 0:
                    cell.fill = FILL_ALT_ROW
        _auto_width(ws_proc)

    wb.save(filename)
    return filename


# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="TidyCPU — CPU Affinity Optimization Utility",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  sudo python3 tidycpu.py                          # View CPU topology only
  sudo python3 tidycpu.py --threads                # Show processes and rebalancing
  sudo python3 tidycpu.py --check-pid 1234         # Check specific process affinity
  sudo python3 tidycpu.py --live                   # Live monitor (5 seconds)
  sudo python3 tidycpu.py --live --duration 10     # Live monitor (10 seconds)
  sudo python3 tidycpu.py --pid 1234               # Monitor single PID
  sudo python3 tidycpu.py --pid nginx              # Monitor process by name
  sudo python3 tidycpu.py --pid "nginx|php-fpm"    # Monitor multiple processes
  sudo python3 tidycpu.py --pid "1234|nginx|5678"  # Mix of PIDs and names
  sudo python3 tidycpu.py --cpu-freq               # Include CPU frequency info
  sudo python3 tidycpu.py --export-html report.html  # Export to HTML
  sudo python3 tidycpu.py --export-excel report.xlsx # Export to Excel (requires openpyxl)
  sudo python3 tidycpu.py --live --ignore-col Bar    # Hide Bar column in live view
  sudo python3 tidycpu.py --live --specify nginx,php-fpm  # Highlight rows for these parents
  sudo python3 tidycpu.py --stack-procs              # Stack all active processes per core row
  sudo python3 tidycpu.py --stack-procs --all        # Include even idle processes in stack
  sudo python3 tidycpu.py --stack-procs --min-usage 0.5  # Only stack processes using >= 0.5% CPU
  sudo python3 tidycpu.py --stack-procs --ignore-process kworker,ksoftirqd  # Exclude kernel threads
        """
    )
    parser.add_argument("--live", "-l", action="store_true",
        help="Live monitoring mode (refreshes every 500ms)")
    parser.add_argument("--duration", "-d", type=int, default=5,
        help="Duration for live mode in seconds (default: 5)")
    parser.add_argument("--pid", "-p", type=str,
        help="PIDs/names to monitor, pipe-separated (e.g. nginx|php-fpm|1234)")
    parser.add_argument("--threads", "-t", action="store_true",
        help="Show thread information for processes")
    parser.add_argument("--cpu-freq", "-f", action="store_true",
        help="Show CPU frequency information (min/max/current)")
    parser.add_argument("--check-pid", type=int, metavar="PID",
        help="Check affinity and CPU usage for a specific process ID")
    parser.add_argument("--export-html", type=str, metavar="FILE",
        help="Export report to HTML file (e.g., report.html)")
    parser.add_argument("--export-text", type=str, metavar="FILE",
        help="Export report to text file (e.g., report.txt)")
    parser.add_argument("--export-excel", type=str, metavar="FILE",
        help="Export report to Excel file (e.g., report.xlsx)  [requires openpyxl]")
    parser.add_argument("--ignore-col", type=lambda s: [c.strip() for c in s.split(",")],
        default=[], metavar="COLS",
        help="Comma-separated columns to hide (e.g. Bar,Usage)")
    parser.add_argument("--specify", type=lambda s: [p.strip() for p in s.split(",")],
        default=None, metavar="PARENTS",
        help="Highlight rows whose parent process matches these names (comma-separated)")
    parser.add_argument("--stack-procs", action="store_true",
        help="Stack all processes running on each core below the core row")
    parser.add_argument("--all", dest="include_all", action="store_true",
        help="Include all processes in the stacked view (even 0%% CPU); use with --stack-procs")
    parser.add_argument("--min-usage", type=float, default=None, metavar="PCT",
        help="Only show processes with CPU%% >= PCT in stacked view (e.g. 0.5); default 0.1 without --all")
    parser.add_argument("--ignore-process", type=lambda s: [p.strip() for p in s.split(",")],
        default=None, metavar="NAMES",
        help="Comma-separated process name prefixes to exclude from stacked view (e.g. kworker,ksoftirqd)")

    args = parser.parse_args()

    # Resolve effective min_usage for --stack-procs:
    #   --all forces 0.0 (show every process, even idle)
    #   --min-usage N overrides the threshold explicitly
    #   default without --all is 0.1 (only threads with measurable recent CPU)
    if args.include_all:
        effective_min_usage = 0.0
    elif args.min_usage is not None:
        effective_min_usage = args.min_usage
    else:
        effective_min_usage = 0.1

    print(BANNER)

    check_root()
    print(f"  {C.GREEN}✔{C.RESET}  Running as root.\n")

    sysinfo = get_system_info(show_cpu_freq=args.cpu_freq)
    print_system_info(sysinfo)

    topology = get_cpu_topology()
    num_cores = len(topology)

    # Resolve --pid (pipe-separated names/PIDs) early so all branches can use it
    resolved_pids: list[int] = []
    if args.pid:
        resolved_pids = resolve_pids(args.pid)

    # ── Check specific PID mode ──────────────────────────────────────────────
    if args.check_pid:
        pid = args.check_pid
        print(f"\n  {C.CYAN}Checking PID {pid}{C.RESET}\n")
        
        rc, out, _ = run(["ps", "-p", str(pid), "-o", "pid,pcpu,comm", "--no-headers"])
        if rc != 0:
            print(f"  {C.RED}✘ PID {pid} not found{C.RESET}\n")
            sys.exit(1)
        
        parts = out.strip().split(None, 2)
        if len(parts) < 3:
            print(f"  {C.RED}✘ Unable to read process info{C.RESET}\n")
            sys.exit(1)
        
        try:
            pid_val = int(parts[0])
            cpu_percent = float(parts[1])
            name = parts[2].strip()
        except ValueError:
            print(f"  {C.RED}✘ Invalid process data{C.RESET}\n")
            sys.exit(1)
        
        rc2, out2, _ = run(["taskset", "-p", str(pid_val)])
        if rc2 != 0:
            print(f"  {C.RED}✘ Unable to read affinity{C.RESET}\n")
            sys.exit(1)
        
        m = re.search(r"current affinity mask:\s*([0-9a-fA-F,]+)", out2)
        if not m:
            print(f"  {C.RED}✘ Unable to parse affinity mask{C.RESET}\n")
            sys.exit(1)
        
        current_cores = mask_to_cores(m.group(1), num_cores)
        core_stats = get_core_usage(sample_ms=500, topology=topology)
        stats_map = {cs.core_id: cs for cs in core_stats}
        
        print(f"{C.BOLD}{'─'*78}{C.RESET}")
        print(f"{C.BOLD}  PROCESS DETAILS{C.RESET}")
        print(f"{C.BOLD}{'─'*78}{C.RESET}")
        print(f"  {C.CYAN}PID:{C.RESET}      {pid_val}")
        print(f"  {C.CYAN}Process:{C.RESET}  {name}")
        print(f"  {C.CYAN}CPU:%{C.RESET}     {C.YELLOW}{cpu_percent:.1f}%{C.RESET}")
        
        cores_str = ",".join(map(str, current_cores)) if current_cores else "all"
        print(f"  {C.CYAN}Cores:{C.RESET}    {cores_str}")
        
        print(f"\n{C.BOLD}  CORE STATUS{C.RESET}")
        print(f"{C.BOLD}{'─'*78}{C.RESET}")
        
        if current_cores:
            for core_id in current_cores:
                if core_id in stats_map:
                    cs = stats_map[core_id]
                    bar = usage_bar(cs.usage, width=20)
                    print(f"  CPU{core_id:>2}  {cs.usage:>5.1f}%  {bar}  {label_color(cs.label)}")
        else:
            print(f"  {C.DIM}Process can run on any core (no affinity set){C.RESET}")
        
        threads = get_threads_for_pid(pid_val, num_cores)
        if threads:
            print(f"\n{C.BOLD}  THREADS ({len(threads)}){C.RESET}")
            print(f"{C.BOLD}{'─'*78}{C.RESET}")
            print(f"  {'TID':>7}  {'Name':<24}  {'CPU%':>6}  Cores")
            print(f"  {'───':>7}  {'────────────────────────':<24}  {'─────':>6}  ─────")
            
            for t in threads[:15]:
                t_cores_str = ",".join(map(str, t.current_cores)) if t.current_cores else "all"
                print(f"  {t.tid:>7}  {t.name:<24}  {C.YELLOW}{t.cpu_percent:>5.1f}%{C.RESET}  {t_cores_str}")
            
            if len(threads) > 15:
                print(f"  {C.DIM}... {len(threads) - 15} more threads{C.RESET}")
        
        print()
        sys.exit(0)

    # ── Live monitor / PID-filter mode ──────────────────────────────────────
    if args.live or resolved_pids:
        try:
            live_monitor(
                duration_sec=args.duration,
                interval_ms=3000,
                filter_pids=resolved_pids or None,
                show_cpu_freq=args.cpu_freq,
                export_html=args.export_html,
                export_text=args.export_text,
                export_excel=args.export_excel,
                sysinfo=sysinfo,
                topology_data=topology,
                ignore_cols=args.ignore_col,
                specify_parents=args.specify,
                stack_procs=args.stack_procs,
                min_usage=effective_min_usage,
                ignore_procs=args.ignore_process,
            )
        except KeyboardInterrupt:
            print(f"\n\n  {C.YELLOW}Monitoring interrupted.{C.RESET}\n")
        return

    # ── Standard rebalance mode ──────────────────────────────────────────────
    print(f"\n  {C.CYAN}○{C.RESET}  Sampling core usage (500 ms) …", end="", flush=True)
    core_stats = get_core_usage(sample_ms=500, topology=topology,
                                stack_procs=args.stack_procs,
                                min_usage=effective_min_usage,
                                ignore_procs=args.ignore_process)
    print(f"\r  {C.GREEN}✔{C.RESET}  Core telemetry collected.          ")

    print_topology(topology, core_stats,
                   ignore_cols=args.ignore_col,
                   specify_parents=args.specify,
                   stack_procs=args.stack_procs)

    processes = []

    if args.threads:
        print(f"\n  {C.CYAN}○{C.RESET}  Identifying top CPU consumers …", end="", flush=True)
        processes = get_top_processes(n=5, num_cores=num_cores, with_threads=True)
        print(f"\r  {C.GREEN}✔{C.RESET}  Process snapshot ready.            ")

    if args.export_html:
        try:
            if not processes:
                processes = get_top_processes(n=5, num_cores=num_cores, with_threads=False)
            filename = export_to_html(sysinfo, topology, core_stats, processes, args.export_html,
                                      ignore_cols=args.ignore_col)
            print(f"\n  {C.GREEN}✔{C.RESET}  HTML report exported to: {C.CYAN}{filename}{C.RESET}")
        except Exception as e:
            print(f"\n  {C.RED}✘{C.RESET}  HTML export failed: {e}")

    if args.export_text:
        try:
            if not processes:
                processes = get_top_processes(n=5, num_cores=num_cores, with_threads=False)
            filename = export_to_text(sysinfo, topology, core_stats, processes, args.export_text,
                                      ignore_cols=args.ignore_col)
            print(f"\n  {C.GREEN}✔{C.RESET}  Text report exported to: {C.CYAN}{filename}{C.RESET}")
        except Exception as e:
            print(f"\n  {C.RED}✘{C.RESET}  Text export failed: {e}")

    if args.export_excel:
        try:
            if not processes:
                processes = get_top_processes(n=5, num_cores=num_cores, with_threads=False)
            filename = export_to_excel(sysinfo, topology, core_stats, processes, args.export_excel,
                                       ignore_cols=args.ignore_col,
                                       stack_procs=args.stack_procs)
            print(f"\n  {C.GREEN}✔{C.RESET}  Excel report exported to: {C.CYAN}{filename}{C.RESET}")
        except Exception as e:
            print(f"\n  {C.RED}✘{C.RESET}  Excel export failed: {e}")

    if args.threads:
        actions = build_rebalance_plan(core_stats, processes)
        print_rebalance_plan(actions)

        if not actions:
            print(f"\n{C.DIM}  Nothing to do. Exiting.{C.RESET}\n")
            sys.exit(0)

        print(f"\n{C.BOLD}  Apply these changes? (y/n): {C.RESET}", end="", flush=True)
        try:
            answer = input().strip().lower()
        except (EOFError, KeyboardInterrupt):
            print(f"\n{C.YELLOW}  Aborted.{C.RESET}\n")
            sys.exit(0)

        if answer != "y":
            print(f"  {C.YELLOW}⚡  No changes applied. Plan saved for manual review.{C.RESET}\n")
            for a in actions:
                to_s = cores_to_cpulist(a.to_cores)
                print(f"     {C.DIM}sudo taskset -pc {to_s} {a.pid}{C.RESET}")
            print()
            sys.exit(0)

        print(f"\n  {C.CYAN}○{C.RESET}  Applying affinity changes …\n")
        results = [apply_action(a) for a in actions]
        print_results(results)
        print(f"\n{C.GREEN}{C.BOLD}  TidyCPU complete.{C.RESET}\n")
    else:
        print(f"\n{C.DIM}  Use --threads to see processes and rebalancing options.{C.RESET}\n")

if __name__ == "__main__":
    main()